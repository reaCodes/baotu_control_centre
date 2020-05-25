import datetime
import io
import os
import subprocess
import time

import openpyxl
import psutil
import pyautogui
import pyperclip


def parse_excel(excel_path):
    excel_content = list()
    try:
        with open(excel_path, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
    except IOError:
        return False

    wb = openpyxl.load_workbook(in_mem_file, read_only=True)
    st = wb.worksheets[0]
    max_row = st.max_row
    for i in range(max_row):
        excel_content.append([])
        program_id = str(i + 1)
        program_path = st.cell(row=i + 1, column=1).value
        config_path = st.cell(row=i + 1, column=2).value
        excel_content[i].append(program_id)
        excel_content[i].append(program_path)
        excel_content[i].append(config_path)
        i = i + 1

    return excel_content


def income_sum(program_path):
    if "BaoTu.exe" not in program_path:
        return 2
    log_path = program_path.split("BaoTu.exe")[0] + "Log\\" + datetime.date.today().strftime('%Y_%m_%d') + ".log"
    try:
        f_log = open(log_path, 'r', encoding='utf-8')
    except IOError:
        return False
    gold = 0
    silver = 0
    for line in f_log.readlines():
        if "金豆" in line:
            i = line.find("豆")
            gold += int(line[i + 1:i + 10].split(' ')[0])
        if "银豆" in line:
            j = line.find("豆")
            silver += int(line[j + 1:j + 10].split(' ')[0])

    return [gold, silver]


def start_prgram(program_path, config_path):
    current_path = os.getcwd()
    if "Task.DouZi.exe" in program_path:
        os.chdir(program_path.split("Task.DouZi.exe")[0])
        program_path = "Task.DouZi.exe"

    process = subprocess.Popen(program_path, creationflags=subprocess.CREATE_NEW_CONSOLE, encoding='utf-8')
    process_pid = process.pid

    time.sleep(1)
    pyperclip.copy(config_path)
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.hotkey('enter')

    os.chdir(current_path)
    return process_pid


def get_db_pi_path(config_path):
    try:
        f_config = open(config_path, 'r', encoding='utf-8')
    except IOError:
        return False

    for line in f_config.readlines():
        if 'HyAccountDb' in line:
            db_path = (line.split('>')[1]).split('<')[0]
        if 'PiHao' in line:
            pi_path = (line.split('>')[1]).split('<')[0]
            break
    f_config.close()

    return [db_path, pi_path]


def pid_is_exist(pid):
    exist = False

    if psutil.pid_exists(pid):
        exist = True
    return exist


def kill(pid):
    if not pid_is_exist(pid):
        return False
    p = psutil.Process(pid)
    p.kill()
    return True


if __name__ == "__main__":
    pass
